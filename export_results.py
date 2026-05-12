import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# Audit results from manual review
RESULTS = [
    ("sample-1153.eml", 43, "Suspicious", "70+",  False, "KetoXplode scam, Reply-To spoofing to winner-win.art"),
    ("sample-120.eml",  6,  "Safe",       "80+",  False, "BTC cryptocurrency giveaway scam"),
    ("sample-1328.eml", 5,  "Safe",       "75+",  False, "Porto Seguro insurance invoice spoofing"),
    ("sample-1458.eml", 8,  "Safe",       "75+",  False, "BB rewards urgency scam"),
    ("sample-1837.eml", 16, "Safe",       "~40",  True,  "Payment decline scare, moderate phishing"),
    ("sample-1941.eml", 17, "Safe",       "~40",  True,  "Amazon order spoofing"),
    ("sample-1977.eml", 11, "Safe",       "80+",  False, "Fake Brazilian legal debt threat"),
    ("sample-2556.eml", 23, "Safe",       "~45",  True,  "Trust Wallet suspension threat"),
    ("sample-266.eml",  15, "Safe",       "~35",  True,  "Suspicious forwarding"),
    ("sample-2836.eml", 10, "Safe",       "70+",  False, "Brazilian rewards liberation scam"),
    ("sample-2941.eml", 21, "Safe",       "~40",  True,  "Auth failures + risky attachment"),
    ("sample-2961.eml", 35, "Suspicious", "~45",  True,  "Portuguese convocation, Gmail sender"),
    ("sample-297.eml",  26, "Safe",       "75+",  False, "MetaMask wallet phishing via PHPMailer"),
    ("sample-3049.eml", 31, "Suspicious", "~55",  True,  "Bradesco Black card spoofing"),
    ("sample-3050.eml", 31, "Suspicious", "~55",  True,  "Bradesco Black spoofing variant"),
    ("sample-3061.eml", 31, "Suspicious", "~55",  True,  "Bradesco spoofing with phishing subject"),
    ("sample-3082.eml", 31, "Suspicious", "~55",  True,  "Bradesco spoofing"),
    ("sample-3084.eml", 31, "Suspicious", "~55",  True,  "Bradesco bank impersonation"),
    ("sample-3140.eml", 31, "Suspicious", "~55",  True,  "Bradesco Black variant"),
    ("sample-3148.eml", 31, "Suspicious", "~55",  True,  "Bradesco Prime variant"),
    ("sample-3286.eml", 5,  "Safe",       "75+",  False, "Porto Seguro invoice spoofing"),
    ("sample-3307.eml", 5,  "Safe",       "75+",  False, "Porto Seguro invoice spoofing variant"),
    ("sample-350.eml",  24, "Safe",       "~40",  True,  "Order hold phishing"),
    ("sample-352.eml",  43, "Suspicious", "~60",  True,  "Microsoft Rewards prize scam"),
    ("sample-3899.eml", 5,  "Safe",       "75+",  False, "Porto Seguro invoice spoofing"),
    ("sample-393.eml",  0,  "Safe",       "80+",  False, "Fake French legal proceeding threat"),
    ("sample-4010.eml", 39, "Suspicious", "75+",  False, "ADAC German auto product scam"),
    ("sample-403.eml",  8,  "Safe",       "75+",  False, "Directaxis loan offer scam"),
    ("sample-434.eml",  15, "Safe",       "80+",  False, "Fake Portuguese police complaint"),
    ("sample-4591.eml", 9,  "Safe",       "75+",  False, "Banking deposit/withdrawal suspension threat"),
    ("sample-542.eml",  14, "Safe",       "80+",  False, "BLUR crypto airdrop scam"),
    ("sample-562.eml",  15, "Safe",       "~35",  True,  "Counterfeit luxury goods marketing"),
    ("sample-6024.eml", 11, "Safe",       "75+",  False, "French credit comparison spoofing"),
    ("sample-6234.eml", 7,  "Safe",       "75+",  False, "Firebase welcome bonus gambling scam"),
    ("sample-6306.eml", 7,  "Safe",       "75+",  False, "Firebase Black Friday bonus scam"),
    ("sample-6363.eml", 12, "Safe",       "75+",  False, "Fake Wellnee knee product scam"),
    ("sample-6382.eml", 21, "Safe",       "75+",  False, "AAA Winter Safety Kit giveaway scam"),
    ("sample-6518.eml", 48, "Suspicious", "~70",  True,  "Substack phishing, manipulated headers"),
    ("sample-6548.eml", 26, "Safe",       "75+",  False, "Wellnee/fake Lidl product scam"),
    ("sample-6560.eml", 26, "Safe",       "75+",  False, "Wellnee variant"),
    ("sample-67.eml",   5,  "Safe",       "75+",  False, "Social media phishing via Google Drive link"),
    ("sample-6831.eml", 29, "Safe",       "~45",  True,  "Dutch energy utility scam (Firebase)"),
    ("sample-7033.eml", 48, "Suspicious", "~65",  True,  "Manipulated headers, HTML obfuscation"),
    ("sample-705.eml",  14, "Safe",       "80+",  False, "+$27,776 balance replenishment scam"),
    ("sample-7116.eml", 24, "Safe",       "~50",  True,  "French identity confirmation phishing"),
    ("sample-713.eml",  19, "Safe",       "80+",  False, "+$19,782 balance cancellation in 24h"),
    ("sample-7204.eml", 19, "Safe",       "75+",  False, "Account unusual traffic restriction scam"),
    ("sample-722.eml",  17, "Safe",       "80+",  False, "+$26,637 balance replenishment scam"),
    ("sample-725.eml",  15, "Safe",       "75+",  False, "Expiring rewards urgency scam"),
    ("sample-726.eml",  15, "Safe",       "75+",  False, "KETO product scam"),
    ("sample-730.eml",  19, "Safe",       "80+",  False, "+$27,755 balance replenishment scam"),
    ("sample-732.eml",  19, "Safe",       "80+",  False, "Balance withdrawal 24h deadline scam"),
    ("sample-735.eml",  19, "Safe",       "80+",  False, "Money withdrawal opportunity scam"),
    ("sample-7364.eml", 16, "Safe",       "80+",  False, "Amazon Prime payment failure with hidden chars"),
    ("sample-766.eml",  14, "Safe",       "80+",  False, "Reward claim prize scam"),
    ("sample-781.eml",  14, "Safe",       "80+",  False, "Hurry up rewards arrived prize scam"),
    ("sample-872.eml",  39, "Suspicious", "~55",  True,  "Europa Park survey offer phishing"),
    ("sample-877.eml",  16, "Safe",       "75+",  False, "Bradesco-Livelo points spoofing"),
    ("sample-886.eml",  50, "Suspicious", "80+",  False, "German fake crime indictment threat"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Audit Results"

# styles
RED_FILL    = PatternFill("solid", fgColor="FF4444")
GREEN_FILL  = PatternFill("solid", fgColor="70AD47")
YELLOW_FILL = PatternFill("solid", fgColor="FFD966")
HEADER_FILL = PatternFill("solid", fgColor="2E4057")
WHITE_FONT  = Font(color="FFFFFF", bold=True)
BOLD        = Font(bold=True)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN        = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

headers = ["File", "Score", "Verdict", "Expected", "Match", "Notes"]
col_widths = [22, 8, 14, 12, 8, 55]

# header row
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill   = HEADER_FILL
    cell.font   = WHITE_FONT
    cell.alignment = CENTER
    cell.border = THIN
    ws.column_dimensions[cell.column_letter].width = w

# data rows
for row_idx, (file, score, verdict, expected, match, notes) in enumerate(RESULTS, 2):
    row_data = [file, score, verdict, expected, "✅" if match else "❌", notes]
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = THIN
        cell.alignment = CENTER if col != 6 else LEFT

    # highlight mismatched rows red
    if not match:
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = RED_FILL
    elif score >= 31:
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = YELLOW_FILL

ws.freeze_panes = "A2"

# summary sheet
ws2 = wb.create_sheet("Summary")
total     = len(RESULTS)
mismatches = sum(1 for *_, match, _ in RESULTS if not match)
correct   = total - mismatches

summary = [
    ("Total emails audited", total),
    ("Correctly scored",     correct),
    ("Underscored (missed phishing)", mismatches),
    ("Accuracy",             f"{correct/total*100:.1f}%"),
]
for r, (label, value) in enumerate(summary, 1):
    ws2.cell(row=r, column=1, value=label).font = BOLD
    ws2.cell(row=r, column=2, value=value)
ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 15

out = Path(__file__).parent / "audit_results.xlsx"
wb.save(out)
print(f"Saved: {out}")
